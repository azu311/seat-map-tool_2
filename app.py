import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill
import io
import re
import copy

st.set_page_config(page_title="座席マップ 青塗りツール", layout="wide")
st.title("🏟️ 座席マップ 青塗りツール")
st.caption("クラス名＋列＋座席番号を入力し、座席シートのセルを青色に塗りつぶします")

# ─────────────────────────────────────────────
# 1. パース関数
# ─────────────────────────────────────────────
def parse_seat_text(text):
    """
    テキストから (class_name, row_num, seat_num) のリストを返す
    例:
      Class S South 1列33
      Class SS End-1 2列　8、9
      Class A Side 3列8,9
    """
    results = []

    # 改行 or 「Class」の前で分割
    # まず全行を「Class」で分割して各エントリを処理
    # 全角スペース→半角
    text = text.replace('\u3000', ' ').replace('　', ' ')

    # Classで始まるブロックに分割
    blocks = re.split(r'(?=Class\s)', text)

    for block in blocks:
        block = block.strip()
        if not block:
            continue

        # クラス名パターン: Class SS-T / Class SS End-1 / Class S South / Class A Side など
        # 列パターン: 数字 + 列
        # 座席パターン: 数字（複数は "、" "," "." で区切り or "25.26" のような形）
        m = re.match(
            r'(Class\s+\S+(?:\s+\S+)?)\s+(\d+)列\s*([\d\s、,．.・]+)',
            block
        )
        if not m:
            # クラス名が3トークンのパターン試行: Class A End-1 など
            m = re.match(
                r'(Class\s+\S+\s+\S+)\s+(\d+)列\s*([\d\s、,．.・]+)',
                block
            )
        if not m:
            continue

        class_name = m.group(1).strip()
        row_num = int(m.group(2))
        seat_str = m.group(3)

        # 座席番号を展開 (区切り文字: 、, ．.)
        seat_parts = re.split(r'[、,．.\s・]+', seat_str.strip())
        for sp in seat_parts:
            sp = sp.strip()
            if sp.isdigit():
                results.append((class_name, row_num, int(sp)))

    # 重複排除
    return list(set(results))


def normalize_class(s):
    """連続スペースを1つにして比較用に正規化"""
    return re.sub(r'\s+', ' ', str(s).strip())


# ─────────────────────────────────────────────
# 2. UI
# ─────────────────────────────────────────────
import datetime

col1, col2 = st.columns([1, 2])

with col1:
    uploaded = st.file_uploader(
        "📂 ベースExcelをアップロード",
        type=["xlsx"],
        help="25－26ブロックマップ_座席番号 / _列 / _クラス シートを含むファイル"
    )
    game_date = st.date_input(
        "📅 試合日付",
        value=datetime.date.today(),
        help="シート名・ファイル名に使用されます（例：1月1日 → 0101）"
    )
    date_str = game_date.strftime("%m%d")
    st.caption(f"シート名・ファイル名に使用される日付コード：**{date_str}**")

with col2:
    seat_text = st.text_area(
        "📝 座席指定テキストを貼り付け",
        height=200,
        placeholder="例:\nClass S South 1列33\nClass S South 1列5\nClass SS End-1 2列8、9"
    )

run = st.button("🎨 青塗り実行", type="primary", disabled=(uploaded is None or not seat_text.strip()))

# ─────────────────────────────────────────────
# 3. 処理
# ─────────────────────────────────────────────
if run:
    with st.spinner("処理中..."):
        # パース
        seats = parse_seat_text(seat_text)
        if not seats:
            st.error("座席指定を解析できませんでした。入力形式を確認してください。")
            st.stop()

        st.write(f"**解析された座席数:** {len(seats)} 件")

        # Excelロード
        wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()))

        required_sheets = ['25－26ブロックマップ_座席番号', '25－26ブロックマップ_列', '25－26ブロックマップ_クラス']
        missing = [s for s in required_sheets if s not in wb.sheetnames]
        if missing:
            st.error(f"必要なシートが見つかりません: {missing}")
            st.stop()

        ws_seat  = wb['25－26ブロックマップ_座席番号']
        ws_row   = wb['25－26ブロックマップ_列']
        ws_class = wb['25－26ブロックマップ_クラス']

        BLUE_FILL = PatternFill("solid", fgColor="0000FF")

        # セル座標マップ構築: (class_name, row_val, seat_val) -> (r, c)
        # 全セルを走査してインデックス化
        coord_map = {}
        max_row = ws_class.max_row
        max_col = ws_class.max_column

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cv = ws_class.cell(row=r, column=c).value
                rv = ws_row.cell(row=r, column=c).value
                sv = ws_seat.cell(row=r, column=c).value

                if cv is None or rv is None or sv is None:
                    continue

                cv_norm = normalize_class(str(cv))
                try:
                    rv_int = int(rv)
                except (ValueError, TypeError):
                    continue
                try:
                    sv_int = int(sv)
                except (ValueError, TypeError):
                    continue

                coord_map[(cv_norm, rv_int, sv_int)] = (r, c)

        # 突合＆塗り
        matched = []
        unmatched = []

        for (class_name, row_num, seat_num) in seats:
            key = (normalize_class(class_name), row_num, seat_num)
            if key in coord_map:
                r, c = coord_map[key]
                ws_seat.cell(row=r, column=c).fill = BLUE_FILL
                matched.append({
                    "クラス": class_name,
                    "列": row_num,
                    "座席": seat_num,
                    "セル": f"R{r}C{c}"
                })
            else:
                unmatched.append({
                    "クラス": class_name,
                    "列": row_num,
                    "座席": seat_num
                })

        # 出力：座席番号シートだけを新しいワークブックにコピーして出力
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from copy import copy

        wb_out = Workbook()
        wb_out.remove(wb_out.active)  # デフォルトシートを削除

        # 座席番号シートをコピー
        ws_src = ws_seat
        ws_dst = wb_out.create_sheet(date_str)  # シート名を日付4桁に

        # セルの値・スタイルをコピー
        for row in ws_src.iter_rows():
            for cell in row:
                new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font      = copy(cell.font)
                    new_cell.border    = copy(cell.border)
                    new_cell.fill      = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection  = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # 列幅・行高をコピー
        for col in ws_src.column_dimensions:
            ws_dst.column_dimensions[col].width = ws_src.column_dimensions[col].width
        for row in ws_src.row_dimensions:
            ws_dst.row_dimensions[row].height = ws_src.row_dimensions[row].height

        # 結合セルをコピー
        for merge in ws_src.merged_cells.ranges:
            ws_dst.merge_cells(str(merge))

        out_buf = io.BytesIO()
        wb_out.save(out_buf)
        out_buf.seek(0)

        original_name = uploaded.name.replace(".xlsx", "")
        out_name = f"{original_name}_{date_str}_blue_marked.xlsx"

    # ─────────────────────────────────────────────
    # 4. 結果表示
    # ─────────────────────────────────────────────
    st.success(f"✅ 完了！ 塗り: {len(matched)}件 / 未一致: {len(unmatched)}件")

    col_a, col_b = st.columns(2)

    with col_a:
        st.subheader(f"✅ 塗れた席（{len(matched)}件）")
        if matched:
            st.dataframe(matched, use_container_width=True)
        else:
            st.info("一致なし")

    with col_b:
        st.subheader(f"❌ 塗れなかった席（{len(unmatched)}件）")
        if unmatched:
            st.dataframe(unmatched, use_container_width=True)
            st.caption("クラス名・列・座席番号がデータに存在しない可能性があります")
        else:
            st.success("すべての座席が一致しました！")

    st.download_button(
        label="⬇️ 出力Excelをダウンロード",
        data=out_buf,
        file_name=out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )

# ─────────────────────────────────────────────
# サイドバー: ヘルプ
# ─────────────────────────────────────────────
with st.sidebar:
    st.header("📖 入力形式")
    st.markdown("""
**基本形式:**
```
クラス名 列数字列 座席番号
```

**使用例:**
```
Class S South 1列33
Class S South 1列5
Class SS End-1 2列8、9
Class A Side 3列8,9
```

**複数座席の区切り文字:**
- 読点: `8、9`
- カンマ: `8,9`
- スペース: `8 9`
- ピリオド: `25.26`

**利用可能なクラス名:**
- Class S South
- Class S Side
- Class S End-1 / End-2
- Class SS Side
- Class SS End-1 / End-2
- Class SS-T
- Class A South
- Class A Side
- Class A End-1 / End-2
- Class B Side
- Class B End-1 / End-2
""")
